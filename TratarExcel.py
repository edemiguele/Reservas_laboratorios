# -*- coding: utf-8 -*-
# O # -*- coding: cp1252 -*-
# O # -*- coding: latin-1 -*-

import unicodedata
from openpyxl import load_workbook
from openpyxl import Workbook
import json
import mysql.connector
import os
from datetime import datetime



config_path = os.path.join(os.path.dirname(__file__), "config.json")
with open(config_path, "r") as config_file:
    config = json.load(config_file)

# Configuración de la conexión
conexion = mysql.connector.connect(
    host=config["db_host"],       # Dirección del servidor
    user=config["db_user"],       # Usuario de la base de datos
    password=config["db_password"],  # Contraseña del usuario
    database=config["db_name"]    # Nombre de la base de datos
)

# Crear un cursor para ejecutar consultas
cursor = conexion.cursor()


# Ruta del archivo CSV
#ruta_csv = 'D:\Proyectos\DatosExcel\cortinasVirtuales.csv'
#Profesor;código asignatura;Nombre asignatura;Titulación;Asignatura completa;Grupo A;Grupo B;Semestre

# Función para normalizar nombres de columnas
def normalizar(texto):
    return ''.join(
        c for c in unicodedata.normalize('NFD', texto)
        if unicodedata.category(c) != 'Mn'
    ).lower().replace(' ', '_')


# Contar claves relacionadas con "Excel"
contador_excel = 0
cursor.execute("TRUNCATE TABLE virtuales")
cursor.execute("TRUNCATE TABLE restricciones")
conexion.commit()

for clave, valor in config.items():
    if "virtuales" in clave.lower():  # Verifica si "excel" está en el nombre de la clave
        # Vaciar la tabla prereservas antes de insertar nuevos datos
        #cursor.execute("TRUNCATE TABLE prereservas")
        #conexion.commit()
        
        contador_excel += 1
        print(f"Clave: {clave}, Valor: {valor}")

        # Ruta del archivo Excel
        ruta_excel = valor

        # Cargar el archivo Excel
        libro = load_workbook(ruta_excel)
        print(libro.sheetnames)
        # Seleccionar la hoja 'Hoja 1'
        hoja = libro['Hoja 1']  # Asegúrate de usar el nombre exacto de la hoja

        # Iterar sobre las filas desde la fila 4
        for fila in hoja.iter_rows(min_row=4, values_only=True):
            # Verificar si todas las celdas de la fila están vacías
            if fila[0] is None:
                break  # Detener el bucle si la fila está vacía
    
            cod_asignatura = fila[1]  
            grupo_a = fila[5]
            grupo_b = fila[6]
    
            if grupo_a is None or grupo_b is None:
                continue
    
            # Imprimir los valores de la fila (opcional para depuración)
            print(f"codAsignatura: {cod_asignatura}, grupoa: {grupo_a}, grupob: {grupo_b}")

            # Preparar la consulta SQL
            consulta = """
                INSERT INTO virtuales (idasignatura, grupo1, grupo2)
                VALUES (%s, %s, %s)
            """

            # Ejecutar la consulta con los valores
            cursor.execute(consulta, (cod_asignatura, grupo_a, grupo_b))
    if "restricciones" in clave.lower():  # Verifica si "excel" está en el nombre de la clave
        contador_excel += 1
        print(f"Clave: {clave}, Valor: {valor}")
        
        # Ruta del archivo Excel
        ruta_excel = valor

        # Cargar el archivo Excel
        libro = load_workbook(ruta_excel)
        print(libro.sheetnames)
        # Seleccionar la hoja 'Hoja 1'
        hoja = libro['Hoja 1']  # Asegúrate de usar el nombre exacto de la hoja

        # Iterar sobre las filas desde la fila 4
        for fila in hoja.iter_rows(min_row=4, values_only=True):
            # Verificar si todas las celdas de la fila están vacías
            if fila[0] is None:
                break  # Detener el bucle si la fila está vacía
    
            cod_asignatura = fila[1]  
            practicas = fila[4] #si practicas = 0 es que es todas las practicas
            opcion1 = fila[5] 
            opcion2 = fila[6]
            opcion3 = fila[7]

            if practicas is "TODAS":
                practicas = 0
            else:
                practicas = practicas.replace("P", "")

             # Preparar la consulta SQL
            consulta = """
                INSERT INTO restricciones (idasignatura, numpractica, codespacio, prioridad)
                VALUES (%s, %s, %s, %s)
            """

            # Determinar el valor de codespacio basado en las opciones
            if opcion1 is not None:
                codespacio = opcion1
                # Imprimir los valores de la fila (opcional para depuración)
                print(f"codAsignatura: {cod_asignatura}, practicas: {practicas}, codespacio: {codespacio}, prioridad: 1")
                # Ejecutar la consulta con los valores
                cursor.execute(consulta, (cod_asignatura, practicas, codespacio, 1))
                if opcion2 is not None:
                    codespacio = opcion2
                    # Imprimir los valores de la fila (opcional para depuración)
                    print(f"codAsignatura: {cod_asignatura}, practicas: {practicas}, codespacio: {codespacio}, prioridad: 2")
                    # Ejecutar la consulta con los valores
                    cursor.execute(consulta, (cod_asignatura, practicas, codespacio, 2))
                    if opcion3 is not None:
                        codespacio = opcion3  
                        # Imprimir los valores de la fila (opcional para depuración)
                        print(f"codAsignatura: {cod_asignatura}, practicas: {practicas}, codespacio: {codespacio}, prioridad: 3")
                        # Ejecutar la consulta con los valores
                        cursor.execute(consulta, (cod_asignatura, practicas, codespacio, 3))

    if "lugares" in clave.lower():  # Verifica si "excel" está en el nombre de la clave
        contador_excel += 1
        print(f"Clave: {clave}, Valor: {valor}")

        # Ruta del archivo Excel
        ruta_excel = valor

        # Cargar el archivo Excel
        libro = load_workbook(ruta_excel)
        print(libro.sheetnames)
        # Seleccionar la hoja 'Hoja 1'
        hoja = libro['Hoja 1']  # Asegúrate de usar el nombre exacto de la hoja

        # Iterar sobre las filas desde la fila 4
        for fila in hoja.iter_rows(min_row=4, values_only=True):
            # Verificar si todas las celdas de la fila están vacías
            if fila[1] is None:
                break  # Detener el bucle si la fila está vacía
    
            cod_asignatura = fila[1]  
            labdiis = fila[4]
            aulacentro = fila[6]
            aulateoria = fila[5]
            manual = fila[7]
            rotada = fila[8]
            sala2 = fila[9]  # Sala 2, si existe
            sala10 = fila[10]  # Sala 10, si existe
            
            print(f"manual: {manual}")
            print("tipo manual: ", type(manual))
            
            if labdiis is True:                
                activo = 1
            else:                
                if aulacentro is True: aulacentro = 1 
                else: aulacentro = 0
                if aulateoria is True: aulateoria = 1
                else: aulateoria = 0
                if aulacentro == 1 or aulateoria == 1:
                    activo = 0
                else: activo=1

            if sala2 in [True, "=TRUE()", "SI", "true", "True"]:
                sala2 = 1
                activo = 0
            elif sala2 in [False, "=FALSE()", "NO", "false", "False"]:
                sala2 = 0

            if sala10 in [True, "=TRUE()", "SI", "true", "True"]:
                sala10 = 1
                activo = 0
            elif sala10 in [False, "=FALSE()", "NO", "false", "False"]:
                sala10 = 0

            # Normaliza el valor a 1 o 0 para la base de datos
            if manual in [True, "=TRUE()", "SI", "true", "True"]:
                manual = 1
            elif manual in [False, "=FALSE()", "NO", "false", "False"]:
                manual = 0
             
            if rotada is None: rotada=0;
               
            # Preparar la consulta SQL
            #Tabla Asignaturas: codAsignatura, nbAsignatura, codArea, acronimo, centro, cuatrimestre, manual, vinculada, activo, aulacentro, aulateoria, rotada
            consulta = """
                update asignaturas set manual=%s, activo=%s, aulacentro=%s, aulateoria=%s, rotada=%s, sala2=%s, sala10=%s
                where idasignatura=%s
            """
            
            print(f"codAsignatura: {cod_asignatura}, labdiis: {labdiis},  manual: {manual}, activo: {activo}, aulacentro: {aulacentro}, aulateoria: {aulateoria}, rotada: {rotada}, sala2: {sala2}, sala10: {sala10}")
            # Ejecutar la consulta con los valores
            cursor.execute(consulta, (manual, activo, aulacentro, aulateoria, rotada, sala2, sala10, cod_asignatura))

   
# Confirmar los cambios en la base de datos
conexion.commit()

# Cerrar el cursor y la conexión
cursor.close()
conexion.close()

