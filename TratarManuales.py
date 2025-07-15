
# -*- coding: utf-8 -*-
# O # -*- coding: cp1252 -*-
# O # -*- coding: latin-1 -*-

import unicodedata
from openpyxl import load_workbook
from openpyxl import Workbook
import json
import mysql.connector
import os
from datetime import datetime


config_path = os.path.join(os.path.dirname(__file__), "config.json")
with open(config_path, "r") as config_file:
    config = json.load(config_file)

# Configuración de la conexión
conexion = mysql.connector.connect(
    host=config["db_host"],       # Dirección del servidor
    user=config["db_user"],       # Usuario de la base de datos
    password=config["db_password"],  # Contraseña del usuario
    database=config["db_name"]    # Nombre de la base de datos
)

# Crear un cursor para ejecutar consultas
cursor = conexion.cursor()


# Ruta del archivo CSV
#ruta_csv = 'D:\Proyectos\DatosExcel\cortinasVirtuales.csv'
#Profesor;código asignatura;Nombre asignatura;Titulación;Asignatura completa;Grupo A;Grupo B;Semestre

# Función para normalizar nombres de columnas
def normalizar(texto):
    return ''.join(
        c for c in unicodedata.normalize('NFD', texto)
        if unicodedata.category(c) != 'Mn'
    ).lower().replace(' ', '_')


# Contar claves relacionadas con "Excel"
contador_excel = 0

for clave, valor in config.items():
    if "manual" in clave.lower():  # Verifica si "excel" está en el nombre de la clave
       
        contador_excel += 1
        print(f"Clave: {clave}, Valor: {valor}")

        # Ruta del archivo Excel
        ruta_excel = valor

        # Cargar el archivo Excel
        libro = load_workbook(ruta_excel)
        print(libro.sheetnames)
        # Seleccionar la hoja 'Hoja 1'
        hoja = libro['Reservas Manual']  # Asegúrate de usar el nombre exacto de la hoja

        # Iterar sobre las filas desde la fila 2
        for fila in hoja.iter_rows(min_row=2, values_only=True):
            # Verificar si todas las celdas de la fila están vacías
            if fila[0] is None:
                break  # Detener el bucle si la fila está vacía
    
            cod_asignatura = fila[0]  
            inicio = fila[1]
            fin = fila[2]
            idgrupo = fila[3]
            idespacio = fila[4]
            numpractica = fila[5]
    
            # Imprimir los valores de la fila (opcional para depuración)
            print(f"codAsignatura: {cod_asignatura}, inicio: {inicio}, fin: {fin}, idgrupo: {idgrupo}, idespacio:{idespacio}, numpractica:{numpractica}")

            # Preparar la consulta SQL
            #Tabla Asignaturas: codAsignatura, nbAsignatura, codArea, acronimo, centro, cuatrimestre, manual, vinculada, activo, aulacentro, aulateoria, rotada
            consulta = """
                update prereservas set idespacio=%s, numpractica=%s, observaciones=%s
                where idasignatura=%s and fechaInicio=%s and fechaFin=%s and idgrupo=%s
            """
            
            # Imprimir los valores de la fila (opcional para depuración)
            print(f"codAsignatura: {cod_asignatura}, inicio: {inicio}, fin: {fin}, idgrupo: {idgrupo}, idespacio:{idespacio}, numpractica:{numpractica}")
            # Ejecutar la consulta con los valores
            cursor.execute(consulta, (idespacio, numpractica, numpractica, cod_asignatura, inicio, fin, idgrupo))
   
# Confirmar los cambios en la base de datos
conexion.commit()

# Cerrar el cursor y la conexión
cursor.close()
conexion.close()